import openpyxl
import xlrd
import datetime
import time
from WorkingDay import work_hours
from openpyxl.styles import PatternFill
import win32com.client


########################
### GLOBAL VARIABLES ###
########################


DEFAULT_BASE_XLS_NAME = '''Site Time Details Report.xlsx'''
DEFAULT_EMPLOYEES_NAME = '''Employees File Name.txt'''
DEFAULT_DST_XLS_NAME = '''New Template.xlsx'''
DEFAULT_SPREADSHEET = 0
RELAUNCH_APPLICATION = '''Excel.Application'''
SOURCE_PATH_NAME_FOR_PROJECT = '''C:/Users/ohrabar/PycharmProjects/jp3_2'''
# Some col titles. Change if needed
DEFAULT_NAME = 'Employee'
DEFAULT_DATE = 'Entry Date'
DEFAULT_PROJ = 'Project Name'
DEFAULT_TASK = 'Project Task'
DEFAULT_TIME = 'Employee Time On Task'

DST_EMPLOYEE_NUMBER_COL = 'A'
DST_EMPLOYEE_NAME_COL = 'B'
DST_PROJECT_NAME_COL = 'C'
DST_TASK_COL = 'D'

# Filled in runtime
DATE_MODE = None
DST_STARTING_ROW = 0
DST_STARTING_COL = 0

global_data = {}
BASE_COLS = {}


###############
### CLASSES ###
###############

class User:
    def __init__(self, name):
        self.name = name
        self.projects = {}
        self.totalhours = 0

    def update(self, project, task, day):
        if project not in self.projects:
            self.projects[project] = Project(project)
        self.projects[project].update(task, day)

class Project:
    def __init__(self, name):
        self.name = name
        self.tasks = {}

    def update(self, task, day):
        if task not in self.tasks:
            self.tasks[task] = Task(task)
        self.tasks[task].days[day.day-1] = day

class Task:
    def __init__(self, name):
        self.name = name
        self.days = []
        for i in range(0, 32):
            self.days.append(None)

#################
### FUNCTIONS ###
#################

def get_from_row(row, data_type):
    return row[BASE_COLS[data_type]].value

def get_day(row):
    # First, get some trash: 43368.0
    raw_value = get_from_row(row, DEFAULT_DATE)
    # Convert into date tuple: (2018, 9, 17, 0, 0, 0)
    day = xlrd.xldate_as_tuple(raw_value, DATE_MODE)
    # Convert into date format and get day
    hours_tmp = get_from_row(row, DEFAULT_TIME)
    # Check for non reported day
    if hours_tmp is '':
        hours = 0
    else:
        hours = int(hours_tmp)
        if hours > 23:
            hours = 23
    return datetime.datetime(day[0], day[1], day[2], hours)

def read_row(row):
    name = get_from_row(row, DEFAULT_NAME)
    project = get_from_row(row, DEFAULT_PROJ)
    task = get_from_row(row, DEFAULT_TASK)
    day = get_day(row)
    print(name, project, ":", task, day.day, day.hour)
    if name not in global_data:
        # If doesn't exist yet, create
        global_data[name] = User(name)
    global_data[name].update(project, task, day)

def find_BASE_COLS(row):
    for i in range(0, len(row)): # A to P lookup
        if row[i].value == '':
            continue
        BASE_COLS[row[i].value] = i

def find_starting_row_col(sheet):
    global DST_STARTING_ROW
    global DST_STARTING_COL

    for i in range(1, 10):
        val = sheet['A'+str(i)].value
        if val == '''#''':
            DST_STARTING_ROW = i
            break

    for i in range(0, 10):
        val = sheet[chr(ord('A')+i) + str(DST_STARTING_ROW)].value
        if isinstance(val, datetime.datetime) and val.day == 1:
            DST_STARTING_COL = i
            break

def get_cell_name(col, line):
    col = col + ord('A')
    if col > ord('Z'):
        return str('A') + str(chr(col - ord('Z') - 1 + ord('A'))) + str(line)
    else:
        return chr(col) + str(line)

######################
### IMPLEMENTATION ###
######################

# Get name of basic file
file_name = input("Base XLS file ('" + DEFAULT_BASE_XLS_NAME + "'):")
if file_name == '':
    file_name = DEFAULT_BASE_XLS_NAME

# Open basic workbook and find all titles/modes
src_wbook = xlrd.open_workbook(file_name)
DATE_MODE = src_wbook.datemode
src_wsheet = src_wbook.sheet_by_index(DEFAULT_SPREADSHEET)
find_BASE_COLS(src_wsheet.row(0))

# Now parse the whole basic file
print('Parsing data... Please wait')
time.sleep(1)
for i in range(1, src_wsheet.nrows):
    if src_wsheet.cell(i, BASE_COLS[DEFAULT_NAME]).value == '':
        continue # Skip empty cells
    read_row(src_wsheet.row(i))
print('\nParsing Done!\n')



# Get list of names to form the report
file_name = input("Base XLS file ('" + DEFAULT_EMPLOYEES_NAME + "'):")
if file_name == '':
    file_name = DEFAULT_EMPLOYEES_NAME

# Find all employees for report
employees_list = []
employees_not_found_list = []
with open(file_name, 'r') as f:
    for line in f:
        tmp = line.split(',')
        for i in tmp:
            i = i.strip()
            if i in global_data:
                employees_list.append(i)
            else:
                employees_not_found_list.append(i)

if len(employees_not_found_list) > 0:
    print('Following employees were not found:')
    print(*employees_not_found_list, sep = ', ')
print('\nPlease wait! Forming report for following employees:')
print(*employees_list, sep = ', ')

###################################################################################
# Open template and refresh date
FileName = DEFAULT_DST_XLS_NAME
# Open Excel
Application = win32com.client.Dispatch(RELAUNCH_APPLICATION)
# Show Excel. While this is not required, it can help with debugging (1-shown file or 0-hidden file)
Application.Visible = 1
# Open Your Workbook
Workbook = Application.Workbooks.open(SOURCE_PATH_NAME_FOR_PROJECT + '/' + FileName)
# Refesh All
Workbook.RefreshAll()
# Saves the Workbook
Workbook.Save()
# Closes Excel
Application.Quit()
##################################################################################

# Get name of destination file
file_name = input("Destination XLS file ('" + DEFAULT_DST_XLS_NAME + "'):")
if file_name == '':
    file_name = DEFAULT_DST_XLS_NAME

# Open destination workbook and find starting col
dst_wbook = openpyxl.load_workbook(file_name, data_only=True)
dst_wsheet = dst_wbook.active
find_starting_row_col(dst_wsheet)

S = dst_wsheet
S['D4'] = "=today()"
S['D3'] = "Manager 1"
print()
# Fill destination XLS
work_hours = 16
S['AM8'] = work_hours
print(work_hours, "======================")
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFD700', end_color='FFFFD700', fill_type='solid')
#whiteFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
line = DST_STARTING_ROW + 1
num = 1
for employee in employees_list:
    emp = global_data[employee]
    S[DST_EMPLOYEE_NUMBER_COL + str(line)] = str(num)
    S[DST_EMPLOYEE_NAME_COL + str(line)] = emp.name
    print("EMPLOYEE:")
    print("---" + employee)
    x = 0
    for tmp, proj in emp.projects.items():
        print("PROJECT:")
        print("------" + proj.name)
        for tmp, task in proj.tasks.items():
            print("TASK:")
            print("---------" + task.name)
            for i in range(0, len(task.days)):
                print(task.days[i])
                if task.days[i] is None:
                    continue
                else:
                    x += task.days[i].hour
    print(x, "====")
    emp.totalhours = x
    if emp.totalhours < work_hours:
        S['AL' + str(line)].fill = redFill
    elif emp.totalhours > work_hours:
        S['AL' + str(line)].fill = yellowFill
    # elif emp.totalhours == work_hours:
    #     S['AL' + str(line)].fill = whiteFill

    S['AL' + str(line)] = emp.totalhours


    for tmp, proj in emp.projects.items():
        S[DST_PROJECT_NAME_COL + str(line)] = proj.name
        for tmp, task in proj.tasks.items():
            S[DST_TASK_COL + str(line)] = task.name
            for i in range(0, len(task.days)):
                if task.days[i] != None:
                    S[get_cell_name(DST_STARTING_COL + i, str(line))] = int(task.days[i].hour)
            # Go to next row

            line = line + 1
    num = num + 1

#***************************** Add value to the "Total time" cell
project_total_row = "AK"
row = 8
#print((S[DST_TASK_COL + str(row)]).value)
while S[DST_TASK_COL + str(row)].value is not None:
    #print(str('+++')+ str(project_total_row))
    S['AK' + str(row)] = str("=SUM(E" + str(row) + ":AJ" + str(row) + ")")
    #print((S['AK' + str(row)]).value)
    row = row + 1
#***********************************************************

print('Saving results into result.xlsx')
dst_wbook.save('result.xlsx')
print('Done!')