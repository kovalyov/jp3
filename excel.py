import datetime

import datetime as datetime
from openpyxl import load_workbook
import re
from WorkingDay import end
import sys

# ************************* Load XLSX file for analyze **************************************#

time = datetime.datetime.now().strftime("_%I_%M_%S_%B_%d_%Y")
# print("Введите название файла :\n", "Пример - C:/Users/%USERNAME%/PycharmProjects/untitled/first_file.xlsx")
# file_name = input() + ".xlsx"
file_name = 'SiteTimeDetailsReport.xlsx'

print("==================")
print(end)
print("==================")

# ************************* Load New Template.xlsx file for saving result *************************#
output = 'reports/NewTemplate.xlsx'
wb_output = load_workbook(output)
sheet_output = wb_output['Timesheet']

# ************************* Load TXT file with EmployeesFileName ****************************#
employeesFileName = 'EmployeesFileName.txt'
file = open(employeesFileName, 'r')
s = file.read()
employees = re.split(r', ', s)
file.close()

# ************************** Load in the workbook ********************************************#
wb = load_workbook(file_name)
# Get sheet names
print(wb.sheetnames)

sheet = wb['Site Time Details Report']
print("Works in sheet: ", str(sheet.title))

ch = 'A'
row = 1

# sheet['T1'] = "Total Time"  # создаем ячейку "Total Time" в которую будем вносить значения общего времени

EmployeeColumn = 0
DateEntryColumn = 0
EmployeeTimeOnTaskColumn = 0
TotalTimeColumn = 0
rows = {}

# ************************* Define columns for work in the file SiteTimeDetailsReport ********#
while chr(ord(ch)) <= "Z":  # Limitation a range of columns
    # print("Column Symbol:", ch)
    cell = sheet[str(ch) + str(row)]
    if cell.value == "Employee":
        EmployeeColumn = cell.column
    if cell.value == "Entry Date":
        DateEntryColumn = cell.column
    if cell.value == "Employee Time On Task":
        EmployeeTimeOnTaskColumn = cell.column
    ch = chr(ord(ch) + 1)
    # row += 1

# print("First_Employee           : ", str(EmployeeColumn))
# print("Second_EntryDate         : ", str(DateEntryColumn))
# print("Third EmployeeTimeOnTask : ", str(EmployeeTimeOnTaskColumn))
# print("Total Time               : ", str(TotalTimeColumn))


# *************************** Sort data by Employees with values ********************************#
for i in employees:
    print("\n")
    print(i)
    ##################################################################
    total_hours = 0
    empty_dates = []
    row_2 = 4
    # print(sheet[str(column_1) + str(row_2)].value)
    while sheet[str(EmployeeColumn) + str(row_2)].value is not None:
        # print(i, '/n')
        # print(sheet[str(column_1) + str(row_2)].value)
        if sheet[str(EmployeeColumn) + str(row_2)].value == i:
            # print(str(sheet[str(column_3) + str(row + 3)].value))

            if sheet[str(EmployeeTimeOnTaskColumn) + str(row_2)].value is None:
                if type(sheet[str(DateEntryColumn) + str(row_2)].value) is not datetime.datetime:
                    # empty_dates.append(str(sheet[str(DateEntryColumn) + str(row_2)].value))
                    formated_day = datetime.datetime.strptime(str(sheet[str(DateEntryColumn) + str(row_2)].value), '%m/%d/%Y')
                    empty_dates.append(formated_day)
                else:
                    empty_dates.append(str(sheet[str(DateEntryColumn) + str(row_2)].value))

            else:
                total_hours += int(sheet[str(EmployeeTimeOnTaskColumn) + str(row_2)].value)
        row_2 += 1
    print("Quantity of hours: ", total_hours)
    if len(empty_dates) != 0:
        print("Empty dates: ", empty_dates)

        # item = empty_dates[0]
        # print(str(item.month))

# *************************** Go to NewTemplate.xlsx and work with it ****************************#
output = 'reports/NewTemplate.xlsx'
wb_output = load_workbook(output)
sheet_output = wb_output['Timesheet']
print("Works in sheet: ", str(sheet_output))
sheet_output["I3"] = end

EmployeeName = 0
ch2 = "A"
row_3 = 7
while chr(ord(ch2)) <= "Z":  # Limitation a range of columns
    cell2 = sheet_output[str(ch2) + str(row_3)]  # try found "Employee Name" cell in the NewTemplate.xlsx file
    if cell2.value == "Employee Name":
        EmployeeName = cell2.column
    ch2 = chr(ord(ch2) + 1)

print("Employee Letter  : ", str(EmployeeName))  # Defined the letter of the column

while sheet_output[str(EmployeeName) + str(row_3)].value is not None:
    row_3 += 1
    for i in employees:
        sheet_output[str(EmployeeName) + str(row_3)].value = i
        row_3 += 1

#################################################################

# while sheet[str(EmployeeColumn) + str(row + 1)].value is not None:
#     a = sheet[str(EmployeeColumn) + str(row + 1)].value
#     b = sheet[str(DateEntryColumn) + str(row + 1)].value
#     res = a + b
#     sheet[str(TotalTimeColumn) + str(row + 1)] = res
#     print(res)
#     row += 1
#


# wb_output.save("reports/NewTemplate" + time + ".xlsx")
